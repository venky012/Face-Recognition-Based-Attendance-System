from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_control
from django.contrib import messages


from accounts.decorators import phone_confirmation_required
from accounts.models import EmployeeProfileInfo
from django.conf import settings


import face_recognition
import cv2
from openpyxl import Workbook, load_workbook
import datetime
import matplotlib.pyplot as plt
import os
# Create your views here.

path = settings.BASE_DIR


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
@phone_confirmation_required
def recogface(request):
    os.environ['DISPLAY'] = ':0'

    # Get a reference to webcam #0 (the default one)
    video_capture = cv2.VideoCapture(0)

    cv2.namedWindow("Video")

    now = datetime.datetime.now()
    today = now.day
    month = now.month 
    year = now.year  

    emp = EmployeeProfileInfo.objects.filter(product_key__product_key=request.user.product_key)
        
    image = []
    ids = []
    names = []

    for i in emp:
        image.append(path+str(i.avatar.url))
        names.append(i.employee_name)
        ids.append(i.employee_id)


    known_face_names = []
    known_face_encodings = []

    warn = []

    j = 0
    for i in image:
        get_image = face_recognition.load_image_file(image[j])
        
        try:
            known_face_encodings.append(face_recognition.face_encodings(get_image)[0])
        except:
            warn.append('face not recognised of employee_id: '+str(ids[j])+' upload a new photo')
            
        known_face_names.append(ids[j])
        j = j + 1

    if len(warn)>0:
        for i in warn:
            messages.warning(request,i)
        # Release handle to the webcam
        video_capture.release()
        cv2.destroyAllWindows()
        return HttpResponseRedirect(reverse('dashboard'))
    
    try:
        book = load_workbook(path+"/facerecog/attendance_sheets/"+str(year)+"_"+str(month)+'.xlsx')
        sheet=book.active
    except:    
        book = Workbook()
        sheet = book.active
        sheet.cell(row = 1, column = 1).value = "Employee name"
        sheet.cell(row = 1, column = 2).value = "Employee Id"
        for i in range(3,34,1):
            sheet.cell(row = 1, column = i).value = i - 2
        j = 2
        for name in names:
            sheet.cell(row = j, column = 1).value = name
            j = j + 1
        j = 2
        for i in ids:
            sheet.cell(row = j, column = 2).value = i
            j = j + 1
        
    # Initialize some variables
    face_locations = []
    face_encodings = []
    face_names = []
    process_this_frame = True
        

        
    
    while True:
    # Grab a single frame of video
        ret, frame = video_capture.read()
        
        # Resize frame of video to 1/4 size for faster face recognition processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        
        # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
        rgb_small_frame = small_frame[:, :, ::-1]
        
        # Only process every other frame of video to save time
        if process_this_frame:
            # Find all the faces and face encodings in the current frame of video
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
        
        face_names = []
        name = "" 
        for face_encoding in face_encodings:
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"
        
            # If a match was found in known_face_encodings, just use the first one.
            if True in matches:
                first_match_index = matches.index(True)
                name = known_face_names[first_match_index]
                # Assign attendance
                if name in ids:
                    
                    index = 2
                    for i in ids:
                        if sheet.cell(row=index,column=2).value == name:
                            print(name)
                            break
                        else:
                            index = index + 1


                    sheet.cell(row=index, column=int(today)+2).value = "Present"
                else:
                    pass
        

        face_names.append(name)
        
        process_this_frame = not process_this_frame
        
        
        # Display the results
        top, right, bottom, left = 0,0,0,0
        for (top, right, bottom, left), name in zip(face_locations, face_names):
            # Scale back up face locations since the frame we detected in was scaled to 1/4 size
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
        
        # Draw a box around the face
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
        
        # Draw a label with a name below the face
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
        
        # Display the resulting image
        cv2.imshow('Video', frame)
        
            
        # Save Woorksheet as present month
        book.save(path+"/facerecog/attendance_sheets/"+str(year)+"_"+str(month)+'.xlsx')
        
        # Hit 'q' on the keyboard to quit!
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
        
    # Release handle to the webcam
    video_capture.release()
    cv2.destroyAllWindows()
    return HttpResponseRedirect(reverse('dashboard'))
